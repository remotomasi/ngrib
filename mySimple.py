from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image
import subprocess

wb = load_workbook("simpleWeatherSimply.xlsx")

ws = wb.active

# calculate total number of rows and 
# columns in source excel file
mr = ws.max_row
mc = ws.max_column

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

blueFill = PatternFill(start_color='FF0000FF',
               end_color='FF0000FF',
               fill_type='solid')
cyanFill = PatternFill(start_color='FF40E0D0',
               end_color='FF40E0D0',
               fill_type='solid')
greenFill = PatternFill(start_color='FF00FF00',
               end_color='FF00FF00',
               fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00',
               end_color='FFFFFF00',
               fill_type='solid')
orangeFill = PatternFill(start_color='FFFF8000',
               end_color='FFFF8000',
               fill_type='solid')
redFill = PatternFill(start_color='FFFF0000',
               end_color='FFFF0000',
               fill_type='solid')
brownFill = PatternFill(start_color='FF964B00',
               end_color='FF964B00',
               fill_type='solid')
darkBrownFill = PatternFill(start_color='FF964BAA',
               end_color='FF964BAA',
               fill_type='solid')


whiteFill = PatternFill(start_color='FFFFFFFF',
               end_color='FFFFFFFF',
               fill_type='solid')
lightSilverFill = PatternFill(start_color='FFE0E0E0',
               end_color='FFE0E0E0',
               fill_type='solid')
silverFill = PatternFill(start_color='FFC0C0C0',
               end_color='FFC0C0C0',
               fill_type='solid')
grayFill = PatternFill(start_color='FFA0A0A0',
               end_color='FFA0A0A0',
               fill_type='solid')
darkGrayFill = PatternFill(start_color='FF808080',
               end_color='FF808080',
               fill_type='solid')
veryGrayFill = PatternFill(start_color='FF505050',
               end_color='FF505050',
               fgColor = 'FFFFFF',
               fill_type='solid')
veryDarkGrayFill = PatternFill(start_color='FF303030',
               fgColor = 'FFFFFF',
               fill_type='solid')
blackFill = PatternFill(start_color='FF000000',
               fgColor = 'FFFFFF',
               fill_type='solid')
vvdgf = Font(color='FFFFFFFF')


verylightGreenFill = PatternFill(start_color='FFCCDDCC',
               end_color='FFCCDDCC',
               fill_type='solid')
lightGreenFill = PatternFill(start_color='FFCCFFCC',
               end_color='FFCCFFCC',
               fill_type='solid')
almostGreenFill = PatternFill(start_color='FF66FF66',
               end_color='FF66FF66',
               fill_type='solid')
darkGreenFill = PatternFill(start_color='FF00DD00',
               end_color='FF00DD00',
               fill_type='solid')
darkGreenFill2 = PatternFill(start_color='FF00CC00',
               end_color='FF00CC00',
               fill_type='solid')
darkGreenFill3 = PatternFill(start_color='FF00AA00',
               end_color='FF00AA00',
               fill_type='solid')
darkGreenFill4 = PatternFill(start_color='FF008800',
               end_color='FF008800',
               fill_type='solid')
darkGreenFill5 = PatternFill(start_color='FF006600',
               end_color='FF006600',
               fill_type='solid')
darkGreenFill6 = PatternFill(start_color='FF004400',
               end_color='FF004400',
               fill_type='solid')
darkGreenFill7 = PatternFill(start_color='FF002200',
               end_color='FF002200',
               fill_type='solid')
darkGreenFill8 = PatternFill(start_color='FF001100',
               end_color='FF001100',
               fill_type='solid')


# writing titles
ws.column_dimensions['A'].width = 21
ws.cell(row = 1, column = 1).value = "Date"
ws.column_dimensions['A'].width = 21
ws.cell(row = 1, column = 2).value = "Temp (°C)"
ws.column_dimensions['B'].width = 10
ws.cell(row = 1, column = 3).value = "Umidita' (%)"
ws.column_dimensions['C'].width = 13
ws.cell(row = 1, column = 4).value = "Vento (km/h)"
ws.column_dimensions['D'].width = 13
ws.cell(row = 1, column = 5).value = "Direzione"
ws.column_dimensions['E'].width = 10
ws.cell(row = 1, column = 6).value = "Pioggia"
ws.column_dimensions['F'].width = 7
ws.cell(row = 1, column = 7).value = "Neve"
ws.column_dimensions['G'].width = 7
ws.cell(row = 1, column = 8).value = "Nuvole basse"
ws.column_dimensions['H'].width = 13
ws.cell(row = 1, column = 9).value = "Nuvole medie"
ws.column_dimensions['I'].width = 13

# formatting date adding the day name
for x in range(2, 46, 1):
    cell = 'A' + str(x)
    dayt = subprocess.run(['date', '--date=' + ws[cell].value + ''], capture_output=True, text=True)
    dayDate = ws[cell].value
    day = print(dayt.stdout[0:3])
    ws[cell].value = ws[cell].value + ' ' + dayt.stdout[0:3]

# formatting titles
for j in range (2, 10):
    if ws.cell(row = 1, column = j): ws.cell(row = 1, column = j).font = Font(color = 'FFFFFFFF', bold = True)
    if ws.cell(row = 1, column = j): ws.cell(row = 1, column = j).fill = PatternFill(start_color = 'FF0000FF', end_color = 'FF0000FF', fill_type = 'solid')

# Temperatura
for x in range(2, 46, 1):
    cell = 'B' + str(x)
    if float(ws[cell].value) >= 273.15: ws[cell].value = float(ws[cell].value) - 273.15
    if float(ws[cell].value) <= 0: ws[cell].fill = blueFill
    if float(ws[cell].value) > 0 and float(ws[cell].value) <= 10: ws[cell].fill = blueFill
    if float(ws[cell].value) >= 11 and float(ws[cell].value) <= 13: ws[cell].fill = cyanFill
    if float(ws[cell].value) > 13 and float(ws[cell].value) <= 22: ws[cell].fill = greenFill
    if float(ws[cell].value) > 22 and float(ws[cell].value) <= 30: ws[cell].fill = yellowFill
    if float(ws[cell].value) > 30 and float(ws[cell].value) <= 35: ws[cell].fill = orangeFill
    if float(ws[cell].value) > 35 and float(ws[cell].value) < 40: ws[cell].fill = redFill
    if float(ws[cell].value) >= 40 and float(ws[cell].value) < 45: 
        ws[cell].fill = brownFill
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 45: 
        ws[cell].fill = darkBrownFill
        ws[cell].font = vvdgf

# Umidita'
for x in range(2, 46, 1):
    cell = 'C' + str(x)
    if float(ws[cell].value) >= 0 and float(ws[cell].value) <= 25: ws[cell].fill = orangeFill
    if float(ws[cell].value) > 25 and float(ws[cell].value) <= 30: ws[cell].fill = yellowFill
    if float(ws[cell].value) > 30 and float(ws[cell].value) <= 60: ws[cell].fill = greenFill
    if float(ws[cell].value) > 60 and float(ws[cell].value) <= 70: ws[cell].fill = cyanFill
    if float(ws[cell].value) > 70 and float(ws[cell].value) <= 100: 
        ws[cell].fill = blueFill
        ws[cell].font = vvdgf

# wind power
for x in range(2, 46, 1):
    cell = 'D' + str(x)
    if float(ws[cell].value) >= 0 and float(ws[cell].value) <= 1: ws[cell].fill = whiteFill
    if float(ws[cell].value) >= 2 and float(ws[cell].value) <= 5: ws[cell].fill = lightGreenFill
    if float(ws[cell].value) >= 6 and float(ws[cell].value) <= 11: ws[cell].fill = almostGreenFill
    if float(ws[cell].value) >= 12 and float(ws[cell].value) <= 19: ws[cell].fill = greenFill
    if float(ws[cell].value) >= 20 and float(ws[cell].value) <= 28: ws[cell].fill = darkGreenFill
    if float(ws[cell].value) >= 29 and float(ws[cell].value) <= 38: ws[cell].fill = darkGreenFill2
    if float(ws[cell].value) >= 39 and float(ws[cell].value) <= 49: 
        ws[cell].fill = darkGreenFill3
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 50 and float(ws[cell].value) <= 61: 
        ws[cell].fill = darkGreenFill4
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 62 and float(ws[cell].value) <= 74: 
        ws[cell].fill = darkGreenFill5
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 75 and float(ws[cell].value) <= 88: 
        ws[cell].fill = darkGreenFill6
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 89 and float(ws[cell].value) <= 102: 
        ws[cell].fill = darkGreenFill7
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 103 and float(ws[cell].value) <= 117: 
        ws[cell].fill = darkGrayFill8
        ws[cell].font = vvdgf
    if float(ws[cell].value) >= 118: ws[cell].fill = blackFill

# delete all the images in the workbook
ws._images = []
for x in range(2, 46, 1):
    cell = 'E' + str(x)
    if float(ws[cell].value) >= 335 or float(ws[cell].value) <= 25: 
        img = Image('icons/n.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if float(ws[cell].value) > 25 and float(ws[cell].value) <= 65: 
        img = Image('icons/ne.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if float(ws[cell].value) > 65 and float(ws[cell].value) <= 115: 
        img = Image('icons/e.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if int(ws[cell].value) > 115 and int(ws[cell].value) <= 155: 
        img = Image('icons/se.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if float(ws[cell].value) > 155 and float(ws[cell].value) <= 205: 
        img = Image('icons/s.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if float(ws[cell].value) > 205 and float(ws[cell].value) <= 245: 
        img = Image('icons/sw.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if float(ws[cell].value) > 245 and float(ws[cell].value) <= 295: 
        img = Image('icons/w.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"
    if float(ws[cell].value) > 295 and float(ws[cell].value) < 335: 
        img = Image('icons/nw.png')
        img.height = 20
        img.width = 20
        ws.add_image(img, cell)
        ws[cell] = "1"

# rain possiblity
for x in range(2, 46, 1):
    cell = 'F' + str(x)
    if float(ws[cell].value) == 1: 
        ws[cell].fill = cyanFill
        ws[cell] = ""
    else: 
        ws[cell].fill = whiteFill
        ws[cell] = ""

# snow possibility
for x in range(2, 46, 1):
    cell = 'G' + str(x)
    if float(ws[cell].value) == 1: 
        ws[cell].fill = greenFill
        ws[cell] = ""
    else: 
        ws[cell].fill = whiteFill
        ws[cell] = ""

# nuvole basse
for x in range(2, 46, 1):
    cell = 'H' + str(x)
    if float(ws[cell].value) > 0 and float(ws[cell].value) <= 5: ws[cell].fill = whiteFill
    if float(ws[cell].value) > 5 and float(ws[cell].value) <= 10: ws[cell].fill = lightSilverFill
    if float(ws[cell].value) > 10 and float(ws[cell].value) <= 30: ws[cell].fill = silverFill
    if float(ws[cell].value) > 22 and float(ws[cell].value) <= 30: ws[cell].fill = grayFill
    if float(ws[cell].value) > 30 and float(ws[cell].value) <= 60: ws[cell].fill = darkGrayFill
    if float(ws[cell].value) > 60 and float(ws[cell].value) <= 80: 
        ws[cell].fill = veryGrayFill
        ws[cell].font = vvdgf
    if float(ws[cell].value) > 80 and float(ws[cell].value) <= 100: 
        ws[cell].fill = veryDarkGrayFill
        ws[cell].font = vvdgf

# nuvole medie
for x in range(2, 46, 1):
    cell = 'I' + str(x)
    if float(ws[cell].value) > 0 and float(ws[cell].value) <= 5: ws[cell].fill = whiteFill
    if float(ws[cell].value) > 5 and float(ws[cell].value) <= 10: ws[cell].fill = lightSilverFill
    if float(ws[cell].value) > 10 and float(ws[cell].value) <= 30: ws[cell].fill = silverFill
    if float(ws[cell].value) > 22 and float(ws[cell].value) <= 30: ws[cell].fill = grayFill
    if float(ws[cell].value) > 30 and float(ws[cell].value) <= 60: ws[cell].fill = darkGrayFill
    if float(ws[cell].value) > 60 and float(ws[cell].value) <= 80: 
        ws[cell].fill = veryGrayFill
        ws[cell].font = vvdgf
    if float(ws[cell].value) > 80 and float(ws[cell].value) <= 100: 
        ws[cell].fill = veryDarkGrayFill
        ws[cell].font = vvdgf

# adding thin borders
for i in range (1, mr + 1):
    for j in range(1, mc + 1):
        ws.cell(row = i, column = j).border = thin_border

wb.save("simpleWeatherSimply.xlsx")
