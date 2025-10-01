# importing openpyxl module
import openpyxl as xl;
# importing math module
import math
  
# opening the source excel file
filename ="final.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file 
filename1 ="data/simpleWeatherSimply.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and 
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# erase the file content
ws2.delete_cols(1, mr)
ws2.delete_rows(1, mc)

# copying the cell values from source 
# excel file to destination excel file

# Date
for i in range (1, mr + 1):
    for j in range(1, 2):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j).value = c.value

# Calculate wind Power and Direction
for i in range (1, mr + 1):
    for j in range(98, 99):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j-95).value = c.value

# RH and calculate wind Power and Direction
for i in range (1, mr + 1):
    for j in range(100, 103):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j-95).value = c.value
        
    if c.value is None:
        c.value = 0.0

        # writing the wind power to destination excel file
        if i > 1 and j == 101: ws2.cell(row = i, column = 13).value = int(math.sqrt(float(c.value) * float(c.value) + float(ws1.cell(row = i, column = 102).value) * float(ws1.cell(row = i, column = 102).value))*3.6)
        if i > 1 and j == 101: ws2.cell(row = i, column = 14).value = int(math.atan2(float(c.value), float(ws1.cell(row = i, column = 102).value))*57.3 + 180)

# CRAIN - CSNOW / probability of rain and snow
for i in range (1, mr + 1):
    for j in range(115,117):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j-105).value = c.value

# Clouds
for i in range (1, mr + 1):
    for j in range(125,128,2):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j-110).value = c.value

# Move columns
ws2.move_range("M1:N45", rows=0, cols=-7)
ws2.move_range("I1:J45", rows=0, cols=-1)
ws2.move_range("L1:L45", rows=0, cols=-2)
ws2.move_range("O1:O45", rows=0, cols=-3)
ws2.move_range("Q1:Q45", rows=0, cols=-4)
ws2.move_range("S1:S45", rows=0, cols=-5)
ws2.move_range("C1:N45", rows=0, cols=-1)

# Rounding and translating from kelvin to celsius temp and dpt
for i in range (2, mr + 1):
    for j in range(2, 3):
        # reading cell value from source excel file
        c = ws2.cell(row = i, column = j)
    
    if c.value is None:
        c.value = 0.0
    
        # writing the wind power to destination excel file
        ws2.cell(row = i, column = j).value = int(float(c.value) - 273.15)

# Rounding rh
for i in range (2, mr + 1):
    for j in range(4, 5):
        # reading cell value from source excel file
        c = ws2.cell(row = i, column = j)
    
    if c.value is None:
        c.value = 0.0

        # writing the wind power to destination excel file
        ws2.cell(row = i, column = j).value = int(float(c.value))

# Move columns
ws2.move_range("D1:F45", rows=0, cols=-1)
ws2.move_range("H1:I45", rows=0, cols=-1)
ws2.move_range("J1:L45", rows=0, cols=-2)
ws2.move_range("G1:J45", rows=0, cols=-1)

# saving the destination excel file
wb2.save(str(filename1))